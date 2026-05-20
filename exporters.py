"""Export scrape results to various formats."""

import csv
import json
import sqlite3
from pathlib import Path
from typing import Any, Optional

from scraper import ScrapeResult


class BaseExporter:
    """Base class for exporters."""

    def export(
        self, result: ScrapeResult, filepath: Optional[str] = None
    ) -> str:
        raise NotImplementedError


class JSONExporter(BaseExporter):
    """Export to pretty-printed JSON."""

    def export(
        self, result: ScrapeResult, filepath: Optional[str] = None
    ) -> str:
        data = result.to_dict()
        content = json.dumps(data, indent=2, default=str, ensure_ascii=False)
        if filepath:
            Path(filepath).write_text(content, encoding="utf-8")
        return content


class CSVExporter(BaseExporter):
    """Export to CSV with proper escaping."""

    def export(
        self, result: ScrapeResult, filepath: Optional[str] = None
    ) -> str:
        rows: list[dict[str, str]] = []
        for key, value in result.data.items():
            if isinstance(value, list):
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        for sub_key, sub_val in item.items():
                            rows.append(
                                {
                                    "section": f"{key}[{i}]",
                                    "field": sub_key,
                                    "value": str(sub_val),
                                }
                            )
                    else:
                        rows.append(
                            {"section": key, "field": str(i), "value": str(item)}
                        )
            elif isinstance(value, dict):
                for sub_key, sub_val in value.items():
                    rows.append(
                        {"section": key, "field": sub_key, "value": str(sub_val)}
                    )
            else:
                rows.append({"section": key, "field": "", "value": str(value)})

        if not rows:
            return ""

        output_path = filepath or "scrape_output.csv"
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["section", "field", "value"])
            writer.writeheader()
            writer.writerows(rows)
        return output_path


class XLSXExporter(BaseExporter):
    """Export to formatted Excel with multiple sheets."""

    def export(
        self, result: ScrapeResult, filepath: Optional[str] = None
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export")

        output_path = filepath or "scrape_output.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")

        for key, value in result.data.items():
            sheet_name = key[:31]
            ws = wb.create_sheet(title=sheet_name)
            if isinstance(value, list) and value and isinstance(value[0], dict):
                headers = list(value[0].keys())
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                for item in value:
                    ws.append([str(item.get(h, "")) for h in headers])
            elif isinstance(value, list):
                ws.append([key])
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                for item in value:
                    ws.append([str(item)])
            elif isinstance(value, dict):
                ws.append(["Key", "Value"])
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                for k, v in value.items():
                    ws.append([k, str(v)])
            else:
                ws.append(["Value"])
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                ws.append([str(value)])

        wb.save(output_path)
        return output_path


class MarkdownExporter(BaseExporter):
    """Export to Markdown for documentation."""

    def export(
        self, result: ScrapeResult, filepath: Optional[str] = None
    ) -> str:
        lines = [f"# Scrape Result: {result.url}\n"]
        lines.append(f"**Content Hash:** `{result.content_hash}`\n")

        for key, value in result.data.items():
            lines.append(f"\n## {key.title()}\n")
            if isinstance(value, list) and value and isinstance(value[0], dict):
                headers = list(value[0].keys())
                lines.append("| " + " | ".join(headers) + " |")
                lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for item in value:
                    lines.append(
                        "| " + " | ".join(str(item.get(h, "")) for h in headers) + " |"
                    )
            elif isinstance(value, list):
                for item in value:
                    lines.append(f"- {item}")
            elif isinstance(value, dict):
                for k, v in value.items():
                    lines.append(f"- **{k}:** {v}")
            else:
                lines.append(str(value))

        content = "\n".join(lines)
        if filepath:
            Path(filepath).write_text(content, encoding="utf-8")
        return content


class SQLiteExporter(BaseExporter):
    """Export to SQLite database for large datasets."""

    def export(
        self, result: ScrapeResult, filepath: Optional[str] = None
    ) -> str:
        db_path = filepath or "scrape_output.db"
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS scrape_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT,
                content_hash TEXT,
                timestamp REAL,
                data TEXT
            )
        """
        )
        cursor.execute(
            """
            INSERT INTO scrape_results (url, content_hash, timestamp, data)
            VALUES (?, ?, ?, ?)
        """,
            (
                result.url,
                result.content_hash,
                result.timestamp,
                json.dumps(result.data, default=str),
            ),
        )
        conn.commit()
        conn.close()
        return db_path


EXPORTERS = {
    "json": JSONExporter,
    "csv": CSVExporter,
    "xlsx": XLSXExporter,
    "md": MarkdownExporter,
    "sqlite": SQLiteExporter,
}


def export_result(
    result: ScrapeResult,
    format: str,
    filepath: Optional[str] = None,
) -> str:
    exporter_cls = EXPORTERS.get(format)
    if not exporter_cls:
        raise ValueError(
            f"Unknown format: {format}. Available: {', '.join(EXPORTERS)}"
        )
    exporter = exporter_cls()
    exported = exporter.export(result, filepath)
    # Return the filepath if one was provided, otherwise return the exported content
    if filepath:
        return filepath
    return exported
